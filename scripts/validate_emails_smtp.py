#!/usr/bin/env python3
"""Validate outreach email addresses without sending a message.

The validator checks syntax, DNS/MX, and the remote SMTP server's RCPT TO
response.  It never issues the SMTP DATA command.  Results are written to a
new CSV so that the source database remains unchanged.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass
import hashlib
from pathlib import Path
import random
import re
import smtplib
import socket
import time
from typing import Iterable, Sequence

import dns.exception
import dns.resolver
from openpyxl import load_workbook


EMAIL_RE = re.compile(
    r"^(?=.{1,254}$)(?=.{1,64}@)[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+"
    r"@(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?\.)+"
    r"[A-Za-z]{2,63}$"
)
ACCEPTED_CODES = {250, 251, 252}


@dataclass(frozen=True)
class ValidationResult:
    email: str
    status: str
    send_allowed: bool
    mx_host: str = ""
    smtp_code: int | None = None
    smtp_message: str = ""
    catch_all: bool | None = None
    checked_at_utc: str = ""


def normalize_email(value: object) -> str:
    return str(value or "").strip().lower()


def syntax_is_valid(email: str) -> bool:
    if not EMAIL_RE.fullmatch(email):
        return False
    try:
        email.rsplit("@", 1)[1].encode("idna")
    except UnicodeError:
        return False
    return True


def decode_reply(message: bytes | str | None) -> str:
    if message is None:
        return ""
    if isinstance(message, bytes):
        return message.decode("utf-8", errors="replace").strip()
    return str(message).strip()


def resolve_mail_hosts(domain: str, timeout: float) -> list[str]:
    resolver = dns.resolver.Resolver(configure=True)
    resolver.timeout = timeout
    resolver.lifetime = timeout
    ascii_domain = domain.encode("idna").decode("ascii")

    try:
        answers = resolver.resolve(ascii_domain, "MX")
        hosts = sorted(
            (int(answer.preference), str(answer.exchange).rstrip("."))
            for answer in answers
        )
        # A single MX record with target "." explicitly rejects email.
        if any(not host for _, host in hosts):
            return []
        return [host for _, host in hosts]
    except dns.resolver.NoAnswer:
        # RFC-compatible fallback for older domains that accept mail on A/AAAA.
        for record_type in ("A", "AAAA"):
            try:
                if resolver.resolve(ascii_domain, record_type):
                    return [ascii_domain]
            except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN, dns.exception.DNSException):
                continue
        return []
    except dns.resolver.NXDOMAIN:
        return []


def classify_rcpt(code: int) -> str:
    if code in ACCEPTED_CODES:
        return "accepted"
    if 400 <= code <= 499:
        return "temporary_or_policy_block"
    if 500 <= code <= 599:
        return "rejected"
    return "unknown"


def smtp_probe(
    email: str,
    mx_hosts: Sequence[str],
    *,
    helo_domain: str,
    mail_from: str,
    timeout: float,
    catch_all_check: bool,
    max_mx_hosts: int,
) -> ValidationResult:
    domain = email.rsplit("@", 1)[1]
    last_error = ""

    for mx_host in mx_hosts[:max_mx_hosts]:
        smtp: smtplib.SMTP | None = None
        try:
            smtp = smtplib.SMTP(timeout=timeout)
            smtp.connect(mx_host, 25)
            code, reply = smtp.ehlo(helo_domain)
            if code >= 400:
                code, reply = smtp.helo(helo_domain)
            if code >= 400:
                last_error = f"HELO rejected: {code} {decode_reply(reply)}"
                continue

            code, reply = smtp.mail(mail_from)
            if code >= 400:
                last_error = f"MAIL FROM rejected: {code} {decode_reply(reply)}"
                continue

            code, reply = smtp.rcpt(email)
            message = decode_reply(reply)
            rcpt_class = classify_rcpt(code)

            if rcpt_class == "rejected":
                return ValidationResult(
                    email=email,
                    status="undeliverable",
                    send_allowed=False,
                    mx_host=mx_host,
                    smtp_code=code,
                    smtp_message=message,
                    catch_all=False,
                )
            if rcpt_class != "accepted":
                return ValidationResult(
                    email=email,
                    status="unknown",
                    send_allowed=False,
                    mx_host=mx_host,
                    smtp_code=code,
                    smtp_message=message,
                )

            if not catch_all_check:
                return ValidationResult(
                    email=email,
                    status="deliverable",
                    send_allowed=True,
                    mx_host=mx_host,
                    smtp_code=code,
                    smtp_message=message,
                    catch_all=None,
                )

            nonce = hashlib.sha256(f"{email}:{time.time_ns()}".encode()).hexdigest()[:20]
            random_recipient = f"adria-check-{nonce}@{domain}"
            random_code, random_reply = smtp.rcpt(random_recipient)
            random_class = classify_rcpt(random_code)

            if random_class == "accepted":
                return ValidationResult(
                    email=email,
                    status="catch_all",
                    send_allowed=False,
                    mx_host=mx_host,
                    smtp_code=code,
                    smtp_message=message,
                    catch_all=True,
                )
            if random_class == "rejected":
                return ValidationResult(
                    email=email,
                    status="deliverable",
                    send_allowed=True,
                    mx_host=mx_host,
                    smtp_code=code,
                    smtp_message=message,
                    catch_all=False,
                )
            return ValidationResult(
                email=email,
                status="accepted_unconfirmed",
                send_allowed=False,
                mx_host=mx_host,
                smtp_code=random_code,
                smtp_message=f"Catch-all test inconclusive: {decode_reply(random_reply)}",
                catch_all=None,
            )
        except (OSError, socket.timeout, smtplib.SMTPException) as exc:
            last_error = f"{type(exc).__name__}: {exc}"
        finally:
            if smtp is not None:
                try:
                    smtp.quit()
                except (OSError, smtplib.SMTPException):
                    smtp.close()

    return ValidationResult(
        email=email,
        status="smtp_unreachable",
        send_allowed=False,
        smtp_message=last_error,
    )


def validate_email(
    email: str,
    *,
    helo_domain: str,
    mail_from: str,
    timeout: float,
    catch_all_check: bool,
    max_mx_hosts: int,
) -> ValidationResult:
    from datetime import datetime, timezone

    checked_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    email = normalize_email(email)
    if not syntax_is_valid(email):
        return ValidationResult(email, "invalid_syntax", False, checked_at_utc=checked_at)

    domain = email.rsplit("@", 1)[1]
    try:
        mx_hosts = resolve_mail_hosts(domain, timeout)
    except dns.exception.DNSException as exc:
        return ValidationResult(
            email,
            "dns_error",
            False,
            smtp_message=f"{type(exc).__name__}: {exc}",
            checked_at_utc=checked_at,
        )
    if not mx_hosts:
        return ValidationResult(email, "no_mx", False, checked_at_utc=checked_at)

    result = smtp_probe(
        email,
        mx_hosts,
        helo_domain=helo_domain,
        mail_from=mail_from,
        timeout=timeout,
        catch_all_check=catch_all_check,
        max_mx_hosts=max_mx_hosts,
    )
    return ValidationResult(**{**asdict(result), "checked_at_utc": checked_at})


def read_rows(path: Path, sheet_name: str, email_column: str) -> list[dict[str, object]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Input must be .xlsx or .csv")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(values)]
    if email_column not in headers:
        raise ValueError(f"Column not found: {email_column}")
    return [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]


def filtered_rows(
    rows: Iterable[dict[str, object]],
    *,
    email_column: str,
    countries: set[str],
    facility_type: str,
    source: str,
) -> list[dict[str, object]]:
    selected = []
    seen_emails = set()
    for row in rows:
        if countries and str(row.get("Country") or "") not in countries:
            continue
        if facility_type and str(row.get("Facility type") or "") != facility_type:
            continue
        row_source = str(row.get("Email_Source") or "")
        if source == "original" and row_source:
            continue
        if source == "inferred" and row_source != "inferred:domain_pattern":
            continue
        email = normalize_email(row.get(email_column))
        if not email or email in seen_emails:
            continue
        seen_emails.add(email)
        selected.append(row)
    return selected


def write_results(
    output_path: Path,
    rows: Sequence[dict[str, object]],
    results: dict[str, ValidationResult],
    email_column: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_columns = list(rows[0].keys()) if rows else [email_column]
    result_columns = [
        "SMTP_Status",
        "Send_Allowed",
        "MX_Host",
        "SMTP_Code",
        "SMTP_Message",
        "Catch_All",
        "Checked_At_UTC",
    ]
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=source_columns + result_columns)
        writer.writeheader()
        for row in rows:
            result = results[normalize_email(row.get(email_column))]
            writer.writerow(
                {
                    **row,
                    "SMTP_Status": result.status,
                    "Send_Allowed": result.send_allowed,
                    "MX_Host": result.mx_host,
                    "SMTP_Code": result.smtp_code or "",
                    "SMTP_Message": result.smtp_message,
                    "Catch_All": "" if result.catch_all is None else result.catch_all,
                    "Checked_At_UTC": result.checked_at_utc,
                }
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--sheet", default="All Facilities")
    parser.add_argument("--email-column", default="Email")
    parser.add_argument("--country", action="append", default=[])
    parser.add_argument("--facility-type", default="")
    parser.add_argument("--source", choices=("all", "original", "inferred"), default="all")
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Maximum addresses checked in one daily batch (default and hard cap: 10)",
    )
    parser.add_argument("--timeout", type=float, default=8.0)
    parser.add_argument("--delay", type=float, default=2.0)
    parser.add_argument("--helo-domain", required=True)
    parser.add_argument("--mail-from", required=True)
    parser.add_argument("--max-mx-hosts", type=int, default=2)
    parser.add_argument("--no-catch-all", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = filtered_rows(
        read_rows(args.input, args.sheet, args.email_column),
        email_column=args.email_column,
        countries=set(args.country),
        facility_type=args.facility_type,
        source=args.source,
    )
    if not 1 <= args.limit <= 10:
        raise ValueError("--limit must be between 1 and the daily cap of 10")
    rows = rows[: args.limit]
    if not rows:
        raise ValueError("No rows matched the requested filters")

    results: dict[str, ValidationResult] = {}
    for index, row in enumerate(rows, start=1):
        email = normalize_email(row.get(args.email_column))
        results[email] = validate_email(
            email,
            helo_domain=args.helo_domain,
            mail_from=args.mail_from,
            timeout=args.timeout,
            catch_all_check=not args.no_catch_all,
            max_mx_hosts=args.max_mx_hosts,
        )
        print(f"[{index}/{len(rows)}] {email.rsplit('@', 1)[-1]}: {results[email].status}")
        if index < len(rows):
            time.sleep(args.delay + random.uniform(0, min(args.delay, 1.0)))

    write_results(args.output, rows, results, args.email_column)
    allowed = sum(result.send_allowed for result in results.values())
    print(f"Saved {len(rows)} results to {args.output}; send allowed: {allowed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
